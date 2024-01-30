import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

# Lee el archivo Excel y crea un DataFrame con los datos
archiv_excel = pd.read_excel('supermarket_sales.xlsx')
# print(archiv_excel[['Gender', 'Product line', 'Total']])

# Crea una tabla pivote a partir del DataFrame
table_pivote = archiv_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
#print(table_pivote)


# Guarda la tabla pivote en un nuevo archivo Excel llamado 'sales_2022.xlsx' en la hoja 'Report' comenzando desde la fila 4
table_pivote.to_excel('sales_2022.xlsx', startrow=4, sheet_name='Report')


# Carga el archivo Excel 'sales_2022.xlsx'
wb = load_workbook('sales_2022.xlsx')
pestania = wb['Report']
# Obtiene las dimensiones de la hoja de cálculo activa
min_col = wb.active.min_column
max_col = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row


# Crea un gráfico de barras
barchart = BarChart()

# Define las referencias de datos y categorías para el gráfico
data = Reference(pestania, min_col=min_col+1, min_row=min_fila, max_col=max_col, max_row=max_fila)
categoria = Reference(pestania, min_col=min_col, min_row=min_fila +1, max_col=min_col, max_row=max_fila)

# Agrega los datos y categorías al gráfico de barras
barchart.add_data(data, titles_from_data=True)
barchart.add_data(categoria)

# Añade el gráfico de barras a la hoja de cálculo en la celda 'B12'
pestania.add_chart(barchart, 'B12')

# Establece el título y el estilo del gráfico de barras
barchart.title = 'Ventas'
barchart.style = 5

# Establece el estilo de las letras en la hoja de cálculo
pestania['A1'] = 'Reporte'
pestania['A2'] = 2022

pestania['A1'].font = Font('Arial', bold=True, size=20)
pestania['A2'].font = Font('Arial', bold=True, size=15)









# Guarda los cambios en el archivo 'sales_2022.xlsx'
wb.save('sales_2022.xlsx')

